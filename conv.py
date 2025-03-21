from openpyxl import load_workbook
from sortedcontainers import SortedSet
import json

workbook = load_workbook(filename="uwu.xlsx", read_only=True)
sheet = workbook.active

headers = ["sbd", "ten", "ngay_sinh", "noi_sinh", "gioi_tinh", "lop", "truong", "mon_thi", "diem", "ket_qua", "xep_giai"]

data = []
for row in sheet.iter_rows():
    row_data = {}
    for header, cell in zip(headers, row):
        if header.strip() != "":
            row_data[header] = cell.value.strip()
    data.append(row_data)

sanitized = {}

for dat in data:
    if dat["diem"] == 'Vắng thi':
        dat["diem"] = 0
    dat["diem"] = float(dat["diem"])
    sanitized[dat["sbd"]] = dat
    del sanitized[dat["sbd"]]["sbd"]

with open("data.json", 'w', encoding='utf-8') as f:
    json.dump(sanitized, f, indent=4, ensure_ascii=False)

workbook.close()